import json

from openpyxl import load_workbook


def write_to_json_file(data, file_path) -> None:
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def parse_menu_excel(file_path) -> list[dict]:
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    menus = []
    current_menu = None
    current_submenu = None

    for row in sheet.iter_rows():
        menu_id = row[0].value if row[0].value else None
        if menu_id:
            menu_title = row[1].value
            menu_description = row[2].value
            current_menu = {'id': menu_id, 'title': menu_title, 'description': menu_description, 'submenus': []}
            menus.append(current_menu)
            current_submenu = None
            continue

        if current_menu is None:  # Skip rows until a menu is defined
            continue

        submenu_id = row[1].value if row[1].value else None
        if submenu_id:
            submenu_title = row[2].value
            submenu_description = row[3].value
            current_submenu = {'id': submenu_id, 'title': submenu_title,
                               'description': submenu_description, 'dishes': []}
            current_menu['submenus'].append(current_submenu)
            continue

        dish_id = row[2].value if row[2].value else None
        if dish_id and current_submenu:
            dish_title = row[3].value
            dish_description = row[4].value
            dish_price = float(row[5].value) if len(row) > 5 and row[5].value else None
            dish_discount = float(row[6].value) if len(row) > 6 and row[6].value else None
            dish = {'id': dish_id, 'title': dish_title, 'description': dish_description,
                    'price': str(dish_price), 'discount': dish_discount}
            current_submenu['dishes'].append(dish)

    return menus
